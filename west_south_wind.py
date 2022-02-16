# -*- coding: utf-8 -*-
"""
Created on Wed Nov 17 16:48:19 2021

@author: user
"""

import os
import numpy as np
from openpyxl import load_workbook, Workbook
import math

#########################################

U_path = '/media/johnny/My Passport/NCDR/Data/test/u_v_data/U_850'
V_path = '/media/johnny/My Passport/NCDR/Data/test/u_v_data/V_850'

time = '12'

width_list = [39,40]

height_list = [23,24]

wind_speed_threshold = 12.86

Result = 'Result/west_south_wind.xlsx'

min_direction = math.cos(math.pi*60/180)

max_direction = math.cos(math.pi*30/180)

##############################################

def dotproduct(v1, v2):
  return sum((a*b) for a, b in zip(v1, v2))

def length(v):
  return math.sqrt(dotproduct(v, v))

def angle(v1, v2):
  return dotproduct(v1, v2) / (length(v1) * length(v2))

print("[INFO] loading U_V data")


wb = Workbook()
sheet = wb.create_sheet("west_south_wind", 0)
count_true = 0
count_false = 0

U_date_list = []
V_date_list = []

U_value_temp = []
V_value_temp = []

U_value_list = []
V_value_list = []



for date in sorted(os.listdir(U_path)):
    #print(file)
    
    U_date = date[:8]
    U_time = date[8:10]
    #print(U_date)
    #print(U_time)
    if U_date not in U_date_list:
        print(U_date)
        U_date_list.append(U_date)
    if U_time == time:
        
        #read file
        path = U_path +  '/' + date
        print(path)
        f = open(path)
        data_U = []
        
        for line in f.readlines():
            
            parsing = line.split()
            
            if parsing[0]=='71':
                continue
            else:
                data_U.append(float(parsing[0]))
                
                
        if not data_U:
            data_U = np.zeros((41,71))
        else:
            data_U = np.array(data_U)
            data_U = np.reshape(data_U,(41,71))
            
        
        
        #get the value in the selected wind field
        for i in width_list:
            for j in height_list:
                U_value_temp.append(data_U[j][i])
                
        U_value_list.append(U_value_temp)
        U_value_temp = []
    
    
                
for date in sorted(os.listdir(V_path)):
    #print(file)
    V_date = date[:-6]
    V_time = date[-6:-4]
    
    if V_date not in V_date_list:
        V_date_list.append(V_date)
    if V_time == time:
        
        #read file
        path = V_path +  '/' + date
        print(path)
        f = open(path)
        data_V = []
        
        for line in f.readlines():
            
            parsing = line.split()
            
            if parsing[0]=='71':
                continue
            else:
                data_V.append(float(parsing[0]))
                
        if not data_V:
            data_V = np.zeros((41,71))
        else:
            data_V = np.array(data_V)
            data_V = np.reshape(data_V,(41,71))
        
        #get the value in the selected wind field
        for i in width_list:
            for j in height_list:
                V_value_temp.append(data_V[j][i])
                
        V_value_list.append(V_value_temp)
        V_value_temp = []
            

if not len(U_value_list) == len(U_date_list):
    print("U vector format not equivalent")
    print("u_value: ",len(U_value_list))
    print('u_data: ',len(U_date_list))
    
if not len(V_value_list) == len(V_date_list):
    print("v vector format not equivalent")
    print("v_value: ",len(V_value_list))
    print('v_data: ',len(V_date_list))
    
if not len(U_date_list) == len(V_date_list):
    print("u,v vector format not equivalent")
    
    
is_wind_speed_ok = 1
is_wind_direction_ok = 1
vector_1 = [1,0]
wind_dir = 0
for i in range(len(U_date_list)):
    
    for j in range(len(height_list)*len(width_list)):
        
        #chek wind speed
        wind_speed = math.sqrt(U_value_list[i][j]*U_value_list[i][j] + V_value_list[i][j]*V_value_list[i][j])
        if wind_speed <wind_speed_threshold:
            #print("speed: ",wind_speed)
            is_wind_speed_ok = 0
            break
        
        wind_dir = angle(vector_1,[U_value_list[i][j],V_value_list[i][j]])
        
        #1 0 -1 0 
        #if wind_dir>0 or wind_dir<-1:
        if not (U_value_list[i][j] > 0 and V_value_list[i][j] > 0 and wind_dir>min_direction and wind_dir<max_direction):
            #print("dir: ",wind_dir)
            is_wind_direction_ok = 0
            break
        #check wind direction
        
    if is_wind_direction_ok and is_wind_speed_ok:
        sheet.append([U_date_list[i],0])
        count_true+=1
        #print(wind_dir)
    else:
        sheet.append([U_date_list[i],1])
        count_false+=1
    
    is_wind_direction_ok = 1
    is_wind_speed_ok = 1


    
                
#print(U_date_list[22])
#print("U: ",U_value_list[22][0])
#print("V: ",V_value_list[22][0])
#print(test)
#print(test_len)
wb.save(Result)
print(count_true)
print(count_false)