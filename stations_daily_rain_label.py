# -*- coding: utf-8 -*-
"""
Created on Mon Nov 15 16:01:22 2021

@author: user
"""

import os
import numpy as np
from openpyxl import load_workbook, Workbook

#########################################

station_path = '/media/ubuntu/My Passport/NCDR/Data/station_data'

Result = 'Result/daily_rain.xlsx'

min_rain = 0

thresh_hold = 40

rate = 0.1


#special_station_input_id = ['C0V250','C0R140']


#south
special_station_input_id = [
    
    '467410','467420','467780','CAN010','CAN020',
    'CAN030','CAN040','CAN050','CAN060','CAN070',
    'CAN080','CAN090','CAN100',
    
    'C0X100','C0X110','C0X310','CM0130','C0X060',
    'C0O860','C0X180','C0X160','C0O840','C0X290',
    'C0X210','C0X020','C0X240','C0X300','C0X200',
    'C0O930','C0X190','C0X150','C0O950','C0X140',
    'C0X080','C0X130','C0X050','C0O830','C0X260',
    'C0X270','C0X320','C0X280','C0X120','C0O900',
    'C0O970','C0O980','C0O910','C0X250','C0O810',
    'C0X220','C0O960','C0O990','C0X170','C0X230',
    
    'C0V250','C0R140'
    ]




#########################################


print("[INFO] loading special station")


wb = Workbook()
sheet = wb.create_sheet("daily_afternoon_rainfall", 0)
count_true = 0
count_false = 0

for year in os.listdir(station_path):
    #print(file)
    year_dir = station_path + "/" + year
    for month in os.listdir(year_dir):
        month_dir = year_dir + "/" + month
        for date in os.listdir(month_dir):
            date_dir = month_dir + "/" + date
            afternoon_rainfall = 0
            stations_daily_rain = {}
            for date_file in os.listdir(date_dir):
                if not date_file.endswith(".txt"):
                    #break
                    continue
                time = int(date_file[-6:-4])
                
                file_name = date_file
                date_txt = date_dir + "/" + date_file
                print(date_txt)
                f = open(date_txt)
                
                station = -1
                
                tmp_water = -1
                
                
                
                for line in f.readlines():
                    line = line.replace("-", " -")
                    parsing = line.split()
                    if parsing[0] not in special_station_input_id:
                        continue
                    else:
                        # parsing append list
                        station = parsing[0]
                        tmp_water= float(parsing[9])
                        if tmp_water<min_rain:
                            tmp_water = min_rain
                        
                        if station in special_station_input_id:
                            if stations_daily_rain.get(station) == None:
                                stations_daily_rain[station] = tmp_water
                            else:
                                stations_daily_rain[station] += stations_daily_rain[station]
                
            all_larger = 0
            
            for station_id in stations_daily_rain:
                if stations_daily_rain[station_id] > thresh_hold:
                    
                    all_larger +=1
                    break
            
            #if all_larger>int(len(stations_daily_rain)*rate):
            print(len(stations_daily_rain))
            if all_larger>0:
                print("1")
                count_true+=1
                sheet.append([date,1])
            else:
                print("0")
                count_false+=1
                sheet.append([date,0])
                
wb.save(Result)
print(count_true)
print(count_false)