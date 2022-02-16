"""
Created on Mon Nov 15 16:01:22 2021

@author: user
"""

import os
import numpy as np
from openpyxl import load_workbook, Workbook
from station_id import north,middle,south

#########################################

station_path = '/media/johnny/My Passport/NCDR/Data/test/station_data'
#station_path = 'E:/tech/ncdr/station_test'

Result = 'Result/afternoon_rain.xlsx'

min_rain = 0

rate = 0.1

special_station_input_id = south

#special_station_input_id = ['C0R140']
'''
#north
north = [
    
    '466940','466950'
    
    'C0B010','C0B020','CAB010','CAB020','CAB030',
    'CAB040','CAB050',
    
    '466920','466930','466910','CAAH60','466960',
    'A0A460',
    
    'C0A9E0','C0A9C0','CM0020','C0A980','C0AH40',
    'C0AC80','C0AI40','C0A9B0','C0AC40','C0AC70',
    'C0A9G0','C0AH70','CAA040','CAA090'
    
    '466880','466900','A0A9M0','466850','A0A9K0',
    
    'C0AD10','C0AD00','C0A9I0','C0A9I1','C0AI30',
    'C0AC60','C0AI00','C0AD40','C0AG80','C0AG90',
    'C0A650','C0AH10','C0A920','C0AI20','C0A640',
    'C0AH00','C0A870','C0A530','C0A540','C0AH50',
    'C0A710','C0A710','C0A930','C0A931','C0A940',
    'C0A570','C0A560','C0AH90','C0A880','C0A970',
    'C0AH80','C0AD20','C0AI10','C0A580','C0ACA0',
    'C0A950','C0A660','C0AH30','C0A860','C0A520',
    'C0A890','C0A550','C0AD30','C0AD50','CAA020',
    'CAA010','CAA030','CAA050','CAA060','CAA070',
    'CAA080'
    
    
    ]



#middle
special_station_input_id = [
    
    '467490','D2F230','467770','CAD010','CAF020',
    'CAF030','CAF040','CAF050','CAF060','CAF070',
    'CAF080','CAF090',
    
    'C0F930','C0F9K0','C0F000','C0F9N0','C0F9X0',
    'C0F990','C0F9A0','C0F970','C0F9Q0','CM0030',
    'C0F0B0','C0F920','C0F9L0','CM0190','CM0040',
    'C0F9T0','C0F0A0','C0F0E0','C0F860','C0F861',
    'C0F9Y0','C0F900','C0F9Z0','C0F0C0','C0F0D0',
    'C0FA00','C0F850','C0F9U0','C0F9S0','C0F9I0',
    'C0F9P0','C0F9V0','C0F9O0','C0F9R0','C0F9M0',
    'CM0010'
    
    
    ]


#south
special_station_input_id = [
    
    '467410','467420','467780','CAN010','CAN020',
    'CAN030','CAN040','CAN050','CAN060','CAN070',
    'CAN080','CAN090','CAN100',
    
    'C0X100','C0X110','C0X310','CM0130','C0X060',
    'C0O860','C0X180','C0X160','C0O840','C0X290',
    'C0X210','C0X020','C0X240','C0X300','C0X200',
    'C0O930','C0X190','C0X150','C0O950','C0X140',
    'C0X080','C0X130','C0X050','C0O830','C0X260',
    'C0X270','C0X320','C0X280','C0X120','C0O900',
    'C0O970','C0O980','C0O910','C0X250','C0O810',
    'C0X220','C0O960','C0O990','C0X170','C0X230',
    
    '467440','467400','CAP010',
    'CAP020','CAP030','CAP040','CAP050','CAP060',
    
    'C0V700','C0V770','C0V730','C0V350','C0V450',
    'C0V680','C0V360','C0V800','C0V810','C0V620',
    'C0V370','C0V250','C0V820','C0V260','C0V150',
    'C0V660','C0V720','C0V530','C0V310','C0V630',
    'C0V790','CM0180','C0V710','C0V210','C0V610',
    'C0V640','C0V490','C0V490','C0V670','C0V750',
    'C0V690','C0V740','C0V500','C0V440','C0V760',
    'C0V400','C0V650',
    
    '467590','467790','CAQ010','CAQ020',
    
    'C0R170','C0R590','C0R100','C0R150','C0R500',
    'C0R650','C1R340','C0R340','C0R720','C0R750',
    'C0R790','C0R830','C0R420','C0R840','C0R590',
    'C0R540','C0R400','C0R670','C0R700','C0R380',
    'C0R660','C0R640','C0R430','CM0160','C0R530',
    'C0R480','C0R580','C0R350','C0R690','C0R360',
    'C0R620','C0R710','C0R730','C0R260','C0R440',
    'C1R440','C0R520','C0R600','C0R720','C0R470',
    'C0R550','C0R560','C0R510','C0R240','C0R190',
    'C0R220','C0R280','C0R370','C0R680','C0R760',
    'C0R780','C0R140','C0R130','C0R570','C0R160',
    'C1R320','C0R770','C0R740','C0R800','C0R810',
    'C0R820'
    ]

'''
#########################################


print("[INFO] loading special station")


wb = Workbook()
sheet = wb.create_sheet("daily_afternoon_rainfall", 0)
count_true = 0
count_false = 0

for year in os.listdir(station_path):
    #print(file)
    year_dir = station_path + "/" + year
    for month in os.listdir(year_dir):
        month_dir = year_dir + "/" + month
        for date in os.listdir(month_dir):
            date_dir = month_dir + "/" + date
            afternoon_rainfall = 0
            stations_morning_rain = {}
            stations_afternoon_rain = {}
            stations_night_rain = {}
            stations_list = []
            for date_file in os.listdir(date_dir):
                if not date_file.endswith(".txt"):
                    #break
                    continue
                time = int(date_file[-6:-4])
                is_morning = 0
                is_afternoon = 0
                is_night = 0
                #morning: 6~11
                if time <12 and time>5:
                    is_morning = 1
                #night: 18~23
                if time <24 and time >17:
                    is_night = 1
                # afternoon: 12~18
                if time <19 and time >11:
                    is_afternoon = 1
                file_name = date_file
                date_txt = date_dir + "/" + date_file
                print(date_txt)
                f = open(date_txt)
                
                station = -1
                
                tmp_water = -1
                
                
                
                for line in f.readlines():
                    line = line.replace("-", " -")
                    parsing = line.split()
                    if parsing[0] not in special_station_input_id:
                        continue
                    else:
                        # parsing append list
                        station = parsing[0]
                        tmp_water= float(parsing[9])
                        if tmp_water<min_rain:
                            tmp_water = min_rain
                        
                        if station in special_station_input_id:
                            
                            if is_morning:
                                if stations_morning_rain.get(station) == None:
                                    stations_morning_rain[station] = tmp_water
                                    if station not in stations_list:
                                        stations_list.append(station)
                                else:
                                    stations_morning_rain[station] += stations_morning_rain[station]
                            if is_afternoon:
                                if stations_afternoon_rain.get(station) == None:
                                    stations_afternoon_rain[station] = tmp_water
                                    if station not in stations_list:
                                        stations_list.append(station)
                                else:
                                    stations_afternoon_rain[station] += stations_afternoon_rain[station]
                            if is_night:
                                if stations_night_rain.get(station) == None:
                                    stations_night_rain[station] = tmp_water
                                    if station not in stations_list:
                                        stations_list.append(station)
                                else:
                                    stations_night_rain[station] += stations_night_rain[station]
                
            all_larger = 0
            
            for station_id in stations_list:
                
                if stations_afternoon_rain[station_id] < stations_morning_rain[station_id] + stations_night_rain[station_id]:
                    
                    all_larger += 1
                    #break
            
            if all_larger>int(len(stations_list)*rate):
                print("1")
                count_true+=1
                sheet.append([date,1])
            else:
                print("0")
                count_false+=1
                sheet.append([date,0])
                
wb.save(Result)
print(count_true)
print(count_false)