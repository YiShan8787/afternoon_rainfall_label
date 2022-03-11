# -*- coding: utf-8 -*-
"""
Created on Wed Nov 17 16:48:19 2021

@author: user
"""

import os
import numpy as np
from openpyxl import load_workbook, Workbook
import cv2
from numpy import interp
import time
import math
from imutils import paths

#########################################


Result = 'Result/front_day_2015.xlsx'

component_size = 700

weather_path = '/media/ubuntu/My Passport/NCDR/Data/test_ablation/weather_image_data'
#weather_path = 'E:/test_weather/2012-2013'

extract_path = '/media/ubuntu/My Passport/NCDR/Data/test_ablation/weather_extract_image/'
#extract_path = 'E:/tech/ncdr/ncdr_image_extract/extract_image/'

#lon and lat for the whole image
lonRange = [88, 160] # flipped from descending to ascending 90 160
latRange = [4,65] # 4 65

taiwan_lat = 23.58
taiwan_lon = 120.58

threshold_distance = 200 #300km

##############################################

def undesired_objects (image):
    img = image.astype('uint8')
    #print("2.5")
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(img, connectivity=4)
    #print("2.6")
    sizes = stats[:, -1]

    #max_label = 1
    #max_size = sizes[1]
    
    label_list = []
    for i in range(2, nb_components):
        if sizes[i] > component_size:
            #max_label = i
            #max_size = sizes[i]
            label_list.append(i)
    
    img2 = np.zeros(output.shape)
    for j in label_list:
        img2[output == j ] = 255
    #np.where(img2,output in label_list,255)
    #cv2.imshow("Biggest component", img2)
    #cv2.waitKey()
    
    return img2

def getDistance (lonA,latA,lonB,latB):
    R = 6371.004  #radius of earth
    C = math.sin(rad(latA)) * math.sin(rad(latB)) + math.cos(rad(latA)) * math.cos(rad(latB)) * math.cos(rad(lonA - lonB))
    return (R * math.acos(C))
    
def rad(d):
    return d * math.pi / 180.0;
    
print("[INFO] loading special station")


wb = Workbook()
sheet = wb.create_sheet("front_day", 0)
count_true = 0
count_false = 0
count_all = 0
imagePaths = list(paths.list_images(weather_path))


for file in sorted(imagePaths):
    #print(file)
    dir_path = os.path.abspath(weather_path)
    #path = dir_path + '\\' + file
    path = file
    file_time = file[-8:-4]
    file_name = file.split('/')[-1]
    
    if file_time != '0000':
        continue
    #print(file[-8:-4])
    date = file_name[3:11]
    print(file_name)
    print(path)
    
    # 讀取圖檔
    img = cv2.imread(path)
    img_hsv=cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    #print("1")
    # lower mask (0-10)
    lower_red = np.array([0,50,50])
    upper_red = np.array([15,255,255])
    mask0 = cv2.inRange(img_hsv, lower_red, upper_red)
    
    # upper mask (170-180)
    lower_red = np.array([160,50,50])
    upper_red = np.array([180,255,255])
    mask1 = cv2.inRange(img_hsv, lower_red, upper_red)
    
    lower_blue = np.array([110,50,50])
    upper_blue = np.array([140,255,255])
    mask3 = cv2.inRange(img_hsv, lower_blue, upper_blue)
    
    # join my masks
    mask = mask0+mask1+mask3
    
    # set my output img to zero everywhere except my mask
    output_img = img.copy()
    output_img[np.where(mask==0)] = 0
    
    # or your HSV image, which I *believe* is what you want
    output_hsv = img_hsv.copy()
    output_hsv[np.where(mask==0)] = 0
    
    #cv2.imshow('result', output_img)
    #cv2.waitKey()
    
    img_gray=cv2.cvtColor(output_img, cv2.COLOR_BGR2GRAY)
    
    #cv2.imshow('gray', img_gray)
    #cv2.waitKey()
    
    kernel = np.ones((3,3), np.uint8)
    dilation = cv2.dilate(img_gray, kernel, iterations = 1)
    
    #cv2.imshow('canny', dilation)
    #cv2.waitKey()
    #print('2')
    biggest_component = undesired_objects(dilation)
    extract_img = img.copy()
    extract_img[np.where(biggest_component==0)] = 0
    
    rows, cols = np.where(biggest_component==255)
    start = time.time()
    dis_low = 999999
    #print("3")
    count_all+=1
    in_flag = 0
    for i in range(len(rows)):
        y_mid = rows[i]
        x_mid = cols[i]
        
        # the range of y and x pixels
        yRange = [0, extract_img.shape[0]]
        xRange = [0, extract_img.shape[1]]
        
        xPixel = x_mid
        yPixel = y_mid
        
        lat = latRange[1] - interp(yPixel, yRange, latRange) # flipped again
        lon = interp(xPixel, xRange, lonRange)
        
        dis = getDistance(taiwan_lon, taiwan_lat, lon, lat)
        if dis<dis_low:
            dis_low = dis
        if dis < threshold_distance:
            print("in")
            count_false+=1
            sheet.append([date,0])
            in_flag = 1
            break
        #print(dis)
        #print("lat: ",lat)
        #print("lon: ",lon)
    if not in_flag:
        sheet.append([date,1])
    print(dis_low)
    end = time.time()
    print(end-start)
    #cv2.imshow('extract', extract_img)
    #cv2.waitKey()
    
    cv2.destroyAllWindows()
    
    
    # 寫入圖檔
    out_path = extract_path + file_name.split('.')[0] + '_extract' +'.' +'jpg'
    print(out_path)
    cv2.imwrite(out_path, extract_img)
                
count_true = count_all-count_false
wb.save(Result)
print(count_true)
print(count_false)