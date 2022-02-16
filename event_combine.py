# -*- coding: utf-8 -*-
"""
Created on Tue Feb  8 12:53:05 2022

@author: user
"""

from openpyxl import load_workbook, Workbook


###################################################

#north 2016
'''
event_files = [
    'Result/daily_rain_n_2016.xlsx',
    'Result/afternoon_rain_n_2016.xlsx',
    'Result/typhoon_day_2016.xlsx',
    'Result/front_day_2016.xlsx',
    'Result/west_south_wind_2016.xlsx'
    ]
'''

#south
event_files = [
    'Result/daily_rain_s_2016.xlsx',
    'Result/afternoon_rain_s_2016.xlsx',
    'Result/typhoon_day_2016.xlsx',
    'Result/front_day_2016.xlsx',
    ]

Result = "Result/event_day.xlsx"


##################################################


time_list = [] 
label_list = []
count_true = 0
count_false =0

for file in event_files:
    wb_label = load_workbook(file)
    
    sheet_label = wb_label.active
    
    tmp_time_list = []
    tmp_label_list = []
    
    for row in sheet_label.rows:
        if str(row[0].value) not in tmp_time_list:
            tmp_time_list.append(str(row[0].value))
            tmp_label_list.append(str(row[1].value))
    
    if not time_list:
        time_list = tmp_time_list
    else:
        if not time_list == tmp_time_list:
            print("day is not equal")
            break
        
    if not label_list:
        label_list = tmp_label_list
    else:
        label_list = [int(a) & int(b) for a,b in zip(label_list,tmp_label_list)]
        
        
wb = Workbook()
sheet = wb.create_sheet("ncdr_event", 0)

for i in range(len(time_list)):
    sheet.append([time_list[i],
                  label_list[i]
                  ])
    if int(label_list[i]) ==0:
        count_false +=1
    else:
        count_true += 1

print(count_true)
print(count_false)
wb.save(Result)