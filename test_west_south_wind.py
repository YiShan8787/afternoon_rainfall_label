# -*- coding: utf-8 -*-
"""
Created on Wed Nov 17 16:48:19 2021

@author: user
"""

import os
import numpy as np
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
import matplotlib as mpl

import geopandas as gpd

#########################################

U_path = 'E:/tech/ncdr/afternoon_rainfall_label/2010072712_U.txt'
V_path = 'E:/tech/ncdr/afternoon_rainfall_label/2010072712_V.txt'

#Result = 'Result/west_south_wind.xlsx'
shape_file_path = 'E:/tech/ncdr/shp/TAIWAN_MASK.shp'


##############################################

def plt_shp(filename):
    
    TWN_CITY = gpd.read_file(shape_file_path)
    long, long2, lat, lat2 = 119.8,122,21.8,25.4
    x = np.arange(long, long2 + 0.01, 71)
    y = np.arange(lat, lat2 + 0.01, 41)
    x, y = np.meshgrid(x, y)

    fig, ax = plt.subplots(figsize=(10,10))
    #bounds = self.bounds
    cmap = mpl.cm.cool
    
    #norm = mpl.colors.BoundaryNorm(bounds, cmap.N)
    norm = mpl.colors.Normalize(vmin=5, vmax=10)
    #plt.contourf(x, y, img[::-1, :], bounds, cmap=self.cmap)
    ax = TWN_CITY.geometry.plot(ax=ax, alpha=0.3)
    plt.xlim(long, long2)
    plt.ylim(lat, lat2)
    plt.xticks(np.arange(120, 122.5, 0.5))
    '''
    plt.colorbar(
        mpl.cm.ScalarMappable(cmap=cmap, norm=norm),
        # cax=ax,
        #boundaries=[-10] + bounds + [100],
        extend='both',
        extendfrac='auto',
        # ticks=bounds,
        spacing='uniform',
    )
    '''
    plt.savefig(filename)
    plt.close()



print("[INFO] loading U_V data")


#wb = Workbook()
#sheet = wb.create_sheet("west_south_wind", 0)
count_true = 0
count_false = 0

U_date_list = []
V_date_list = []

f = open(U_path)
data_U = []

for line in f.readlines():
    
    parsing = line.split()
    
    if parsing[0]=='71':
        continue
    else:
        data_U.append(float(parsing[0]))
        
data_U = np.array(data_U)
data_U = np.reshape(data_U,(41,71))


f = open(V_path)
data_V = []
    
for line in f.readlines():
    
    parsing = line.split()
    
    if parsing[0]=='71':
        continue
    else:
        data_V.append(float(parsing[0]))
                
        
data_V = np.array(data_V)
data_V = np.reshape(data_V,(41,71))

x,y = np.meshgrid(np.linspace(1,71,71),np.linspace(1,41,41))

fig, ax = plt.subplots(figsize=(20,20))

plt.quiver(x,y,data_U,data_V)
plt.show()

#plt_shp('test')

TWN_CITY = gpd.read_file(shape_file_path)
long, long2, lat, lat2 = 119.8,122,21.8,25.4
x = np.arange(long, long2 + 0.01, 71)
y = np.arange(lat, lat2 + 0.01, 41)
x, y = np.meshgrid(x, y)

fig, ax = plt.subplots(figsize=(20,20))
#bounds = self.bounds
cmap = mpl.cm.cool

#norm = mpl.colors.BoundaryNorm(bounds, cmap.N)
norm = mpl.colors.Normalize(vmin=5, vmax=10)
#plt.contourf(x, y, img[::-1, :], bounds, cmap=self.cmap)
ax = TWN_CITY.geometry.plot(ax=ax, alpha=0.3)
plt.xlim(long, long2)
plt.ylim(lat, lat2)
plt.xticks(np.arange(120, 122.5, 0.5))
'''
plt.colorbar(
    mpl.cm.ScalarMappable(cmap=cmap, norm=norm),
    # cax=ax,
    #boundaries=[-10] + bounds + [100],
    extend='both',
    extendfrac='auto',
    # ticks=bounds,
    spacing='uniform',
)
'''

lon = list(np.linspace(100, 135, 71))
lat = list(np.linspace(10, 30, 41))

ax.quiver(lon, lat, data_U, data_V, color='deepskyblue', scale=50,width=0.01)

plt.savefig("test")
#plt.close()


        
#wb.save(Result)
print(data_U[23][39])
print(data_V[23][39])

#plt.quiver(x,y,data_U,data_V)
plt.show()